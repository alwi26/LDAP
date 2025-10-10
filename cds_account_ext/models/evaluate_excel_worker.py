import sys
import io
import json
import gc
from openpyxl import Workbook
from xlcalculator import ModelCompiler, Evaluator
import calendar
from datetime import datetime


def unwrap_value(val):
    try:
        if hasattr(val, "value"):
            return unwrap_value(val.value)
        if isinstance(val, (list, tuple)):
            return [unwrap_value(v) for v in val]
        if isinstance(val, dict):
            return {k: unwrap_value(v) for k, v in val.items()}
        return val
    except Exception:
        return val


def evaluate_json_excel(json_data):
    stream = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, cells in json_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for cell, value in cells.items():
            ws[cell] = value
    wb.save(stream)
    stream.seek(0)

    model = ModelCompiler().read_and_parse_archive(stream)
    evaluator = Evaluator(model)

    result = {}
    for sheet_name, cells in json_data.items():
        result[sheet_name] = {}
        for cell, value in cells.items():
            try:
                if isinstance(value, str) and value.startswith("="):
                    val = evaluator.evaluate(f"{sheet_name}!{cell}")
                    val = unwrap_value(val)
                else:
                    val = value
            except Exception as e:
                val = f"#ERR {e}"
            result[sheet_name][cell] = {"content": val}

    # Bersihkan memori agar worker tidak leak
    del evaluator
    del model
    del wb
    del stream
    gc.collect()

    return result


def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python evaluate_excel_worker.py <input_json_file>", file=sys.stderr
        )
        sys.exit(1)

    json_file = sys.argv[1]
    with open(json_file, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    result = evaluate_json_excel(json_data)
    # konversi semua hasil ke tipe JSON-safe
    print(json.dumps(unwrap_value(result)))


if __name__ == "__main__":
    main()
