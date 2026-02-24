from odoo import _, fields, models, api
import subprocess, json, tempfile
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from xlcalculator import ModelCompiler, Evaluator, xltypes
import io
import base64
import string
import copy
import gc
import os
import sys
import calendar


class SpreadsheetDashboardGroup(models.Model):
    _inherit = "spreadsheet.dashboard.group"

    period_type = fields.Selection(
        [
            ("month", "Month"),
            ("quarter", "Quarter"),
            ("year", "Year"),
        ],
        string="Default Period Type",
    )
    cds_dashboard_date = fields.Date(string="Dashboard Date Start", tracking=True)
    cds_dashboard_date_end = fields.Date(string="Dashboard Date End", tracking=True)
    cds_status = fields.Many2one("cds.report.status", string="Status")
    dashboard_ids = fields.One2many(copy=True)

    def calculate(self):
        for group in self:
            if group.cds_dashboard_date and group.cds_dashboard_date_end:
                for dashboard in group.dashboard_ids:
                    dashboard.write(
                        {
                            "cds_dashboard_date": group.cds_dashboard_date,
                            "cds_dashboard_date_end": group.cds_dashboard_date_end,
                        }
                    )


class SpreadsheetDashboard(models.Model):
    _name = "spreadsheet.dashboard"
    _inherit = ["spreadsheet.dashboard", "mail.thread", "mail.activity.mixin"]

    name = fields.Char(required=True, translate=True, tracking=True)
    dashboard_group_id = fields.Many2one("spreadsheet.dashboard.group", tracking=True)
    cds_json_converted = fields.Char(string="JSON Converted", copy=False)
    cds_dashboard_date = fields.Date(string="Dashboard Date Start", tracking=True)
    cds_dashboard_date_end = fields.Date(string="Dashboard Date End", tracking=True)
    cds_status = fields.Many2one("cds.report.status", string="Status")
    cds_sdic_ids = fields.One2many(
        "spreadsheet.dashboard.ifrs.calculator",
        "spreadsheet_dashboard_id",
        "IFRS Calculator",
        copy=True,
    )

    join3_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 3 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join4_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 4 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join5_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 5 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join6_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 6 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join7_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 7 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join8_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 8 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join9_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 9 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join10_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 10 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )
    join11_matching_field_id = fields.Many2one(
        "ir.model.fields",
        "Join 11 Journal Item Matching Fields",
        domain="[('model_id.model', '=', 'account.move.line')]",
        tracking=True,
    )

    def action_open_dashboard(self):
        self.ensure_one()
        url = "/odoo/dashboards?dashboard_id=%s" % self.id
        return {
            "type": "ir.actions.act_url",
            "url": url,
            "target": "new",
        }

    def generate_all_balance(self):
        self = self.sudo()

        self.message_post(
            body=_("Calculated IFRS Calculator"),
            subtype_xmlid="mail.mt_note",
        )

        for sd in self:
            for sdic in sd.cds_sdic_ids:
                sdic.generate_balance()

    def unwrap_value(self, val):
        try:
            # xlcalculator v0.4+ pakai .value
            if hasattr(val, "value"):
                return val.value
            # kadang list of values
            if isinstance(val, (list, tuple)):
                return [self.unwrap_value(v) for v in val]
            # dictionary / set
            if isinstance(val, dict):
                return {k: self.unwrap_value(v) for k, v in val.items()}
            return val
        except Exception:
            return val

    def evaluate_json_excel(self, json_data):
        tmp_dir = tempfile.gettempdir()
        json_file_path = os.path.join(tmp_dir, "evaluate_excel_input.json")
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f)

        # Path ke script worker di folder yang sama
        script_path = os.path.join(
            os.path.dirname(__file__), "evaluate_excel_worker.py"
        )

        # Gunakan interpreter Python yang sama dengan Odoo
        python_exec = sys.executable
        cmd = [python_exec, script_path, json_file_path]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=1000)

        if result.returncode != 0:
            raise Exception(f"Worker failed: {result.stderr}")

        try:
            return json.loads(result.stdout)
        except Exception:
            return {"error": result.stdout}

    # def evaluate_json_excel(self, json_data):
    #     try:
    #         # 1. Simpan JSON ke Excel di memory (BytesIO)
    #         stream = io.BytesIO()
    #         wb = Workbook()
    #         wb.remove(wb.active)  # hapus sheet default

    #         for sheet_name, cells in json_data.items():
    #             ws = wb.create_sheet(title=sheet_name)
    #             for cell, value in cells.items():
    #                 ws[cell] = value
    #         wb.save(stream)
    #         stream.seek(0)

    #         # 2. Compile sekali saja
    #         model = ModelCompiler().read_and_parse_archive(stream)
    #         evaluator = Evaluator(model)

    #         # 3. Evaluasi semua cell
    #         result = {}
    #         for sheet_name, cells in json_data.items():
    #             result[sheet_name] = {}
    #             for cell, value in cells.items():
    #                 try:
    #                     if isinstance(value, str) and value.startswith("="):
    #                         # formula → evaluasi
    #                         val = evaluator.evaluate(f"{sheet_name}!{cell}")
    #                         val = self.unwrap_value(val)
    #                     else:
    #                         # plain value
    #                         val = value
    #                 except Exception as e:
    #                     val = f"#ERR {e}"
    #                 result[sheet_name][cell] = {"content": val}

    #         return result

    #     finally:
    #         # Bersihkan memory besar agar Odoo worker tidak overload
    #         try:
    #             wb.close()
    #         except Exception:
    #             pass
    #         del wb, model, evaluator, stream, json_data, sheet_name, cells, cell, value
    #         gc.collect()  # Force garbage collection untuk bebaskan RAM

    def convert_json_text(self):
        for dashboard in self:
            # Make sure the calculation is correct
            dashboard.generate_all_balance()

            binary_data = dashboard.spreadsheet_binary_data
            if not binary_data:
                continue

            decoded_bytes = base64.b64decode(binary_data)
            json_data_text = json.loads(decoded_bytes.decode("utf-8"))
            all_lines = dashboard.cds_sdic_ids
            sheets_text = json_data_text.get("sheets")[1]
            cells_text = sheets_text.get("cells", {})

            headers = self.env["cds_pre_account.move.line"].get_header(cells_text)

            # clear cell value
            cells_text.clear()
            # rewrite the cells
            for i, field in enumerate(headers):
                field_obj = self.env[
                    "spreadsheet.dashboard.ifrs.calculator"
                ]._fields.get(field)
                if not field_obj:
                    continue

                field_label = field_obj.string
                letter = string.ascii_uppercase[i]
                no = 1

                # Keep Odoo format
                cells_text[f"{letter}{no}"] = {"content": f"{field_label}"}

                for ifrs in all_lines:
                    no += 1

                    value = getattr(ifrs, field, "")
                    if field_obj.type == "many2one":
                        value = value.display_name if value else ""
                    elif field_obj.type in ("one2many", "many2many"):
                        value = ", ".join(value.mapped("display_name"))
                    value = value or ""

                    cells_text[f"{letter}{no}"] = {"content": f"{value}"}

            # Convert amount formula
            sheet_backup = copy.deepcopy(json_data_text.get("sheets")[0])
            # Backup Data Formula to sheet 3
            backup_name = (
                "Convert " + dashboard.cds_status.cds_name
                or "" + sheet_backup.get("name")
            )
            sheet_backup.update({"name": backup_name})

            # convert value without {"content": 'value'} inside cells
            formula_cell = json_data_text.get("sheets")[0].get("cells")
            data = {
                json_data_text.get("sheets")[0].get("name"): {
                    cell: val["content"] for cell, val in formula_cell.items()
                },
                json_data_text.get("sheets")[1].get("name"): {
                    cell: val["content"] for cell, val in cells_text.items()
                },
            }
            result = self.evaluate_json_excel(data)
            cells_converted = result.get(json_data_text.get("sheets")[0].get("name"))
            cells_backup = sheet_backup.get("cells", {})
            cells_backup.clear()
            cells_backup.update(cells_converted)
            json_data_text.get("sheets").append(sheet_backup)

            check_domain = next(iter(json_data_text.get("lists")), None)
            json_data_text["lists"]["1"] = json_data_text.get("lists").pop(check_domain)
            domain = json_data_text["lists"]["1"]["domain"]
            json_data_text.get("sheets")[1].update({"cells": cells_text})
            json_data_text["lists"]["1"].update({"domain": domain})
            json_data_text["lists"]["1"].update(
                {"model": "spreadsheet.dashboard.ifrs.calculator"}
            )
            dashboard.cds_json_converted = str(json_data_text)


class SpreadsheetDashboardIFRSCalculator(models.Model):
    _name = "spreadsheet.dashboard.ifrs.calculator"

    spreadsheet_dashboard_id = fields.Many2one("spreadsheet.dashboard")
    name = fields.Char("Name", compute="_get_name")
    account_code = fields.Many2one("account.account")
    join3 = fields.Char("Join 3")
    join4 = fields.Char("Join 4")
    join5 = fields.Char("Join 5")
    join6 = fields.Char("Join 6")
    join7 = fields.Char("Join 7")
    join8 = fields.Char("Join 8")
    join9 = fields.Char("Join 9")
    join10 = fields.Char("Join 10")
    join11 = fields.Char("Join 11")
    reportline = fields.Char("Report Line")
    column = fields.Char("Column")
    balance = fields.Monetary(
        string="Balance", currency_field="currency_id", readonly=True
    )
    negative_bool = fields.Boolean("Negative")
    currency_id = fields.Many2one(
        comodel_name="res.currency",
        string="Currency",
        compute="_compute_currency_id",
        store=True,
        readonly=False,
        precompute=True,
        required=True,
    )

    @api.depends(
        "spreadsheet_dashboard_id.cds_dashboard_date", "spreadsheet_dashboard_id.name"
    )
    def _get_name(self):
        for line in self:
            line.name = line.spreadsheet_dashboard_id.name
            if line.spreadsheet_dashboard_id.cds_dashboard_date:
                line.name = (
                    line.spreadsheet_dashboard_id.name
                    + " ["
                    + line.spreadsheet_dashboard_id.cds_dashboard_date.strftime(
                        "%Y-%m-%d"
                    )
                    + "]"
                )
                if line.spreadsheet_dashboard_id.cds_dashboard_date_end:
                    line.name = (
                        line.spreadsheet_dashboard_id.name
                        + " ["
                        + line.spreadsheet_dashboard_id.cds_dashboard_date.strftime(
                            "%Y-%m-%d"
                        )
                        + "-"
                        + line.spreadsheet_dashboard_id.cds_dashboard_date_end.strftime(
                            "%Y-%m-%d"
                        )
                        + "]"
                    )

    @api.depends("spreadsheet_dashboard_id.company_id.currency_id")
    def _compute_currency_id(self):
        for line in self:
            if line.spreadsheet_dashboard_id.company_id.currency_id:
                line.currency_id = (
                    line.spreadsheet_dashboard_id.company_id.currency_id.id
                )
            else:
                line.currency_id = 12

    def generate_balance(self):
        for line in self:
            domain = [
                ("parent_state", "=", "posted"),
                (
                    "move_id.cds_status",
                    "=",
                    line.spreadsheet_dashboard_id.cds_status.id,
                ),
                ("account_id", "=", line.account_code.id),
                ("date", ">=", line.spreadsheet_dashboard_id.cds_dashboard_date),
                ("date", "<=", line.spreadsheet_dashboard_id.cds_dashboard_date_end),
            ]
            if line.join3 and line.spreadsheet_dashboard_id.join3_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join3_matching_field_id.name,
                        "=",
                        line.join3,
                    )
                ]
            if line.join4 and line.spreadsheet_dashboard_id.join4_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join4_matching_field_id.name,
                        "=",
                        line.join4,
                    )
                ]
            if line.join5 and line.spreadsheet_dashboard_id.join5_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join5_matching_field_id.name,
                        "=",
                        line.join5,
                    )
                ]
            if line.join6 and line.spreadsheet_dashboard_id.join6_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join6_matching_field_id.name,
                        "=",
                        line.join6,
                    )
                ]
            if line.join7 and line.spreadsheet_dashboard_id.join7_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join7_matching_field_id.name,
                        "=",
                        line.join7,
                    )
                ]
            if line.join8 and line.spreadsheet_dashboard_id.join8_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join8_matching_field_id.name,
                        "=",
                        line.join8,
                    )
                ]
            if line.join9 and line.spreadsheet_dashboard_id.join9_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join9_matching_field_id.name,
                        "=",
                        line.join9,
                    )
                ]
            if line.join10 and line.spreadsheet_dashboard_id.join10_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join10_matching_field_id.name,
                        "=",
                        line.join10,
                    )
                ]
            if line.join11 and line.spreadsheet_dashboard_id.join11_matching_field_id:
                domain += [
                    (
                        line.spreadsheet_dashboard_id.join11_matching_field_id.name,
                        "=",
                        line.join11,
                    )
                ]
            move_line_ids = self.env["account.move.line"].search(domain)
            balance = sum(move_line.balance for move_line in move_line_ids)
            if line.negative_bool:
                line.balance = balance * -1
            else:
                line.balance = balance
